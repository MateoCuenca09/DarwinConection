import requests
import pandas as pd
import os
from datetime import datetime, timedelta
from tablas import tabla1,tabla2,Tabla3,Tabla4
from obtener_datos import timeouts,encuesta,reclamo,menuAceso,menuPrincipal,start,DNI_inc,derivadas
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PBI import mainPBI


def convertir_fecha(fecha):
    fecha_objeto = datetime.strptime(fecha, "%Y/%m/%d")
    fecha_formateada = fecha_objeto.strftime("%Y-%m-%d")
    return fecha_formateada

def guardar_datos(df,startDate,endDate):
    try:
        # Crear Excel!
        sDate = convertir_fecha(startDate)
        eDate = convertir_fecha(endDate)
        nombre_archivo = f"{sDate}---{eDate}.xlsx"
        ruta_completa = os.path.join('Datos_procesados', nombre_archivo)
        df.to_excel(ruta_completa, index=True)
        print("Datos guardados en Excel correctamente.")
        return ruta_completa
    except Exception as e:
        print("Error al guardar los datos en Excel:", e)

def mod_datos(datos):
    try:
        df = pd.DataFrame(datos)
        Starts = start(df)
        Timeouts = timeouts(df)
        DNI_Inc = DNI_inc(df)
        MenuPrincipal = menuPrincipal(df)
        Derivadas = derivadas(df)
        MenuAceso = menuAceso(df)
        Reclamos = reclamo(df)
        Encuesta = encuesta(df)
        return {"Starts":Starts, "Timeouts":Timeouts, "DNI_Inc":DNI_Inc, 'MenuPrincipal':MenuPrincipal, 
                'Derivadas':Derivadas, 'MenuAceso':MenuAceso, 'Reclamos':Reclamos, 'Encuesta':Encuesta}
    except Exception as e:
        print("Error al procesar datos", e)

def descargar_datos_desde_api(url, token, startDate, endDate):
    try:
        headers = {
            "Authorization": f"Bearer {token}"
        }
        params = {
            "startDate": startDate,
            "endDate": endDate
        }
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        datos = response.json()
        print("Descarga exitosa!")
        return datos

    except requests.exceptions.RequestException as e:
        print("Error al descargar los datos:", e)
        return None

def crear_tablas(datos):
    try:
        starts = datos['Starts']
        DNI_inc = datos['DNI_Inc']
        menuPrincipal = datos['MenuPrincipal']
        timeouts = datos['Timeouts']
        derivadas = datos['Derivadas']
        menuAceso = datos['MenuAceso']
        reclamos = datos['Reclamos']
        encuesta = datos['Encuesta']

        df = tabla1(starts,DNI_inc,menuPrincipal,timeouts,derivadas)
        df1 = tabla2(menuAceso,menuPrincipal)
        df2 = Tabla3(reclamos,starts)
        df3 = Tabla4(encuesta,starts)

        nuevo_df = pd.concat([df, df1, df2, df3])

        print("Tablas y DF creados con exito.")
        return nuevo_df
    except Exception as e:
        print("Error al Tablas y DF:", e)

def main():
    print("Inicia Programa!!!")
    # URL de la API que deseas conectar
    url_api = "https://api.botdarwin.com/data/conversations"

    # Token de autenticación requerido por la API
    token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJjbGllbnQiOiJlZGlzdXIiLCJyb3V0ZXMiOlsiL2FnZW50cyIsIi9pbnRlbnRzIiwiL2RhdGEiXSwiaWF0IjoxNjg5Nzk1MDk1fQ.Iv4iye5d0uLJz2kTOu9GoEFGGtQGtGCyh-tn0EGCbP8"

    # Obtener la fecha actual
    endDate = datetime.now().strftime("%Y/%m/%d")

    # Obtener la fecha de 7 días atrás
    startDate = (datetime.now() - timedelta(days=15)).strftime("%Y/%m/%d")

    datos = descargar_datos_desde_api(url_api, token, startDate, endDate)

    mainPBI(datos)

    datos_procesados = mod_datos(datos)

    df_completo = crear_tablas(datos_procesados)


    ruta_archivo = guardar_datos(df_completo,startDate,endDate)

    embellecer(ruta_archivo)
    
    print("Finaliza Programa!!!")

def embellecer(archivo):
  try:
      wb = load_workbook(archivo)
      # Obtener la hoja "Sheet1"
      hoja = wb['Sheet1']

      # Crear un objeto Font para aplicar formato en negrita
      font_negrita = Font(bold=True)

      # Aplicar formato en negrita a la celda A2
      hoja['C3'].font = font_negrita
      hoja['C4'].font = font_negrita
      hoja['C6'].font = font_negrita
      hoja['C17'].font = font_negrita
      hoja['C22'].font = font_negrita

      hoja.column_dimensions['A'].width = 47
      hoja.column_dimensions['C'].width = 20

      # Establecer la alineación al centro en el rango especificado
      alignment = Alignment(horizontal='center', vertical='center')

      # Establecer el fondo gris claro en dos columnas por 5 filas
      fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
      rangos = [range(2,7), range(8,16), range(17,21), range(22,28)]
      for rango in rangos:
        for fila in rango:  # Rango de filas a aplicar el formato
              hoja.cell(row=fila, column=2).fill = fill  # Columna 1
              hoja.cell(row=fila, column=2).alignment = alignment
              hoja.cell(row=fila, column=3).fill = fill  # Columna 1
              hoja.cell(row=fila, column=3).alignment = alignment
      wb.save(archivo)
      print("Excel modificado con exito!")

  except Exception as e:
      print("Error al embellecer tabla:", e)
